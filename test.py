import datetime
import os
import shutil
from contextlib import suppress
from copy import copy
from pathlib import Path
from time import sleep

import pandas as pd
from openpyxl import load_workbook

from config import tg_token, chat_id, logger, ecp_paths, template_path, owa_username, owa_password, months_normal, saving_path, smtp_host, smtp_author, homebank_login, homebank_password, ip_address
from tools.net_use import net_use
from tools.smtp import smtp_send
from tools.tg import tg_send
from utils.homebank import homebank, check_homebank_and_collection
from utils.odines import odines_part, odines_check_with_collection
from utils.ofd import ofd_distributor
from utils.sprut_cashbook import open_cashbook


def create_collection_file(file_path, cur_day):
    print('CURDAY:', cur_day)
    current_month: int = int(cur_day.split('.')[1])
    current_year: int = int(cur_day.split('.')[2])
    current_month_name = months_normal[current_month]

    main_working_file = None

    for item in os.listdir(saving_path):

        if current_month_name in str(item).lower() and str(current_year) in str(item):

            if "~$" in item:
                item = item.replace("~$", "")

            main_working_file = os.path.join(saving_path, item)

            break

    if not main_working_file:
        # * If there is no file related to current month
        file_name = f"Файл сбора {current_month_name.capitalize()} {current_year}.xlsx"
        main_working_file = os.path.join(saving_path, file_name)

        shutil.copy(template_path, main_working_file)

    collection_file = load_workbook(main_working_file)
    collection_sheet = collection_file['Файл сбора']
    print(f'Main Excel File: {main_working_file}')
    cols_dict = {
        'A': 'Компания',
        'B': 'Дата чека',
        'C': 'Дата и время чека',
        'D': 'Сумма с НДС',
        'E': 'Ерау',
        'F': '1с',
        'G': 'офд',
        'H': 'примечание',
        'I': '',
        'J': 'Номер чека',
        'K': 'Серийный № фиск.регистратора',
        'L': 'Клиент',
        'M': 'Дата создания записи',
        'N': 'Состояние розничного чека'
    }

    df = pd.read_excel(file_path)

    if df.columns[0] != 'Компания':
        df = pd.read_excel(file_path, header=1)

    for i, row in df.iterrows():

        last_row = collection_sheet.max_row + 1

        for col_key, col_name in cols_dict.items():

            source_cell = collection_sheet.cell(row=last_row - 1, column=collection_sheet[col_key + '1'].column)
            new_cell = collection_sheet.cell(row=last_row, column=collection_sheet[col_key + '1'].column)

            new_cell._style = copy(source_cell._style)
            new_cell.font = copy(source_cell.font)
            new_cell.border = copy(source_cell.border)
            new_cell.alignment = copy(source_cell.alignment)

            cell = collection_sheet[f'{col_key}{last_row}']
            try:
                print('#1', row[col_name])
                cell.value = row[col_name]
                cell.alignment = copy(source_cell.alignment)
            except:
                cell.value = None

    collection_file.save(main_working_file)

    return main_working_file


if __name__ == '__main__':

    months = {
        '172.20.1.24': [2, 3, 4],
        '10.70.2.11': [1, 2, 3, 4, 5, 6],
        '10.70.2.2': [2, 7],
        '10.70.2.9': [3, 12],
        '10.70.2.19': [4, 5],
        '10.70.2.10': [4, 6]
    }

    for month in [10, 11, 12, 1, 2, 3, 4, 5, 6, 7]:
        year = 2022 if month >= 10 else 2023
        if month == 12:
            max_days = datetime.datetime(year + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            max_days = datetime.datetime(year, month + 1, 1) - datetime.timedelta(days=1)

        if month not in months.get(ip_address):
            continue
        print(month, max_days.day)
        # continue

        for days in range(1, max_days.day + 1):
            try:
                if month == 11 and days < 18:
                    continue
                # today = datetime.datetime.today().strftime('%d.%m.%Y')
                # today1 = datetime.datetime.today().strftime('%d.%m.%y')

                if days < 10:
                    if month < 10:
                        today = f'0{days}.0{month}.{year}'
                        today1 = f'0{days}.0{month}.{str(year)[:-2]}'
                    else:
                        today = f'0{days}.{month}.{year}'
                        today1 = f'0{days}.{month}.{str(year)[:-2]}'
                else:
                    if month < 10:
                        today = f'{days}.0{month}.{year}'
                        today1 = f'{days}.0{month}.{str(year)[:-2]}'
                    else:
                        today = f'{days}.{month}.{year}'
                        today1 = f'{days}.{month}.{str(year)[:-2]}'

                print('Start day', today)

                # logger.warning(f'Started processing {day}')

                # today = datetime.datetime.now().date()

                day_ = int(today.split('.')[0])
                month_ = int(today.split('.')[1])
                year_ = int(today.split('.')[2])

                today = datetime.date(year_, month_, day_)

                cashbook_day = (today - datetime.timedelta(days=5)).strftime('%d.%m.%Y')

                days = []

                for i in range(7, 1, -1):
                    day = (today - datetime.timedelta(days=i)).strftime('%d.%m.%Y')
                    days.append(day)

                today = today.strftime('%d.%m.%Y')
                # logger.warning(today, cashbook_day)
                print(cashbook_day)
                print(days)
                print('==========================\n')

                # continue
                # days.append('11.08.2023')
                # days = ['12.08.2023']
                logger.info(days)

                # net_use(Path(template_path).parent, owa_username, owa_password)
                # net_use(ecp_paths, owa_username, owa_password)

                tg_send(f'Робот запустился - <b>{today} | {ip_address}</b>\n\nДата для выгрузки чеков из Спрута - <b>{cashbook_day}</b>\n\nДата проверки в 1С - <b>{days}</b>', bot_token=tg_token, chat_id=chat_id)

                if True:

                    # * ----- 1 -----
                    try:
                        filepath = open_cashbook(cashbook_day)
                    except:
                        logger.warning(f"{days} - Пусто в Розничных чеках за {cashbook_day}")
                        continue
                    # filepath = filepath.replace('Documents', 'Downloads') # If you are compiling for the virtual machines

                    # * ----- 2 -----
                    main_file = create_collection_file(filepath, today)
                    Path(filepath).unlink()

                    # # * ----- 3 -----
                    # logger.warning('Начали Epay')
                    logger.info('Начали Epay')
                    for tries in range(5):
                        with suppress(Exception):
                            filepath = homebank(homebank_login, homebank_password, days[0], days[-1])
                            break
                    # main_file = r'\\vault.magnum.local\Common\Stuff\_06_Бухгалтерия\Для робота\Процесс Сверка ОПТа'
                    check_homebank_and_collection(filepath, main_file)
                    Path(filepath).unlink()

                    # # * ----- 4 -----
                    # logger.warning('Начали 1C')
                    # logger.info('Начали 1C')
                    for tries in range(5):
                        if True:

                            all_days = odines_part(days, month_)

                            odines_check_with_collection(all_days, main_file)
                            break

                        # except Exception as err:
                        #     print("ERROR:", err)
                        #     logger.warning(f"ERROR OCCURED: {err}")

                    # * ----- 5 -----
                    # logger.warning('Начали ОФД')
                    logger.info('Начали ОФД')

                    for tries in range(5):
                        with suppress(Exception):
                            ofd_distributor(main_file)
                            break

                    # smtp_send(fr"""Добрый день!
                    # Сверка ОПТа за {today} завершилась успешно, файл сбора лежит в папке {main_file}""",
                    #           to=['Abdykarim.D@magnum.kz', 'Sagimbayeva@magnum.kz', 'Ashirbayeva@magnum.kz'],
                    #           subject=f'Сверка ОПТа за {today}', username=smtp_author, url=smtp_host)

                    logger.warning(f'Законичили отработку за {today} на машине {ip_address}')

                # except Exception as error:
                    # smtp_send(fr"""Добрый день!
                    #                Сверка ОПТа за {today} - ОШИБКА!!!""",
                    #           to=['Abdykarim.D@magnum.kz', 'Mukhtarova@magnum.kz'],
                    #           subject=f'ОШИБКА Сверка ОПТа за {today}', username=smtp_author, url=smtp_host)
                    # tg_send(f'Возникла ошибка - {error}', bot_token=tg_token, chat_id=chat_id)
                    # raise error
            except Exception as e:
                logger.warning(f'Ошибка на машине {ip_address}: {str(e)}')
