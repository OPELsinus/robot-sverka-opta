import datetime
import os
import shutil
import traceback
from contextlib import suppress
from copy import copy
from pathlib import Path
from time import sleep

import pandas as pd
from openpyxl import load_workbook

from config import tg_token, chat_id, logger, ecp_paths, template_path, owa_username, owa_password, months_normal, saving_path, smtp_host, smtp_author, homebank_login, homebank_password, ip_address, email_to, process_day
from tools.net_use import net_use
from tools.smtp import smtp_send
from tools.tg import tg_send
from utils.homebank import homebank, check_homebank_and_collection
from utils.odines import odines_part, odines_check_with_collection
from utils.ofd import ofd_distributor
from utils.sprut_cashbook import open_cashbook


def create_collection_file(file_path, cur_day, bonuses):
    current_month: int = int(cur_day.split('.')[1])
    current_year: int = int(cur_day.split('.')[2])
    current_month_name = months_normal[current_month]

    main_working_file = None

    for item in os.listdir(saving_path):

        if current_month_name in str(item).lower() and str(current_year) in str(item):

            if "~$" in item:
                item = item.replace("~$", "")

            main_working_file = os.path.join(saving_path, item)

            break

    if not main_working_file:
        # * If there is no file related to current month
        file_name = f"Файл сбора {current_month_name.capitalize()} {current_year}.xlsx"
        main_working_file = os.path.join(saving_path, file_name)

        shutil.copy(template_path, main_working_file)

    collection_file = load_workbook(main_working_file)
    collection_sheet = collection_file['Файл сбора']
    print(f'Main Excel File: {main_working_file}')
    cols_dict = {
        'A': 'Компания',
        'B': 'Дата чека',
        'C': 'Дата и время чека',
        'D': 'Сумма с НДС',
        'E': 'Ерау',
        'F': '1с',
        'G': 'офд',
        'H': 'примечание',
        'I': '',
        'J': 'Номер чека',
        'K': 'Серийный № фиск.регистратора',
        'L': 'Клиент',
        'M': 'Дата создания записи',
        'N': 'Состояние розничного чека'
    }

    df = pd.read_excel(file_path)

    if df.columns[0] != 'Компания':
        df = pd.read_excel(file_path, header=1)

    count = 0
    for i, row in df.iterrows():

        last_row = collection_sheet.max_row + 1

        for col_key, col_name in cols_dict.items():

            source_cell = collection_sheet.cell(row=last_row - 1, column=collection_sheet[col_key + '1'].column)
            new_cell = collection_sheet.cell(row=last_row, column=collection_sheet[col_key + '1'].column)

            new_cell._style = copy(source_cell._style)
            new_cell.font = copy(source_cell.font)
            new_cell.border = copy(source_cell.border)
            new_cell.alignment = copy(source_cell.alignment)

            cell = collection_sheet[f'{col_key}{last_row}']
            try:
                print('#1', row[col_name])
                cell.value = row[col_name]
                cell.alignment = copy(source_cell.alignment)
            except:
                cell.value = None

        collection_sheet[f'I{last_row}'].value = bonuses[count]
        count += 1

    collection_file.save(main_working_file)

    return main_working_file


if __name__ == '__main__':

    try:
        print(f'|{ip_address}|')

        # rng = 9
        #
        # if ip_address == '10.70.2.9':
        #     rng = 11
        # if ip_address == '10.70.2.11':
        #     rng = 10
        # rng = 12  # * ----
        #
        # print(rng)
        #
        # for month in [rng]:

        # max_days = datetime.datetime(2023, (month + 1) % 12, 1) - datetime.timedelta(days=1)

        # rng_day = 1
        #
        # if month == 9:
        #     rng_day = 29
        # if month == 10:
        #     rng_day = 1
        # if month == 11:
        #     rng_day = 15
        # if month == 12:
        #     rng_day = 7
        month = 2
        for days in range(1):

            today = datetime.datetime.today().strftime('%d.%m.%Y')
            today1 = datetime.datetime.today().strftime('%d.%m.%y')

            if process_day != '':
                today = process_day
                today1 = f"{process_day.split('.')[0]}.{process_day.split('.')[1]}.{process_day.split('.')[2][2:]}"

            # if days < 10:
            #     if month < 10:
            #         today = f'0{days}.0{month}.2024'
            #         today1 = f'0{days}.0{month}.24'
            #     else:
            #         today = f'0{days}.{month}.2024'
            #         today1 = f'0{days}.{month}.24'
            # else:
            #     if month < 10:
            #         today = f'{days}.0{month}.2024'
            #         today1 = f'{days}.0{month}.24'
            #     else:
            #         today = f'{days}.{month}.2024'
            #         today1 = f'{days}.{month}.24'

            calendar = pd.read_excel(f'\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-sverka-opta\\Производственный календарь 20{today1[-2:]}.xlsx')
            print(calendar[calendar['Day'] == today1])
            cur_day_index = calendar[calendar['Day'] == today1]['Type'].index[0]
            cur_day_type = calendar[calendar['Day'] == today1]['Type'].iloc[0]
            cur_weekday = calendar[calendar['Day'] == today1]['Weekday'].iloc[0]

            if cur_day_type != 'Holiday':
                # print('Started current date: ', yesterday2)
                _ = f"{today1.split('.')[0]}.{today1.split('.')[1]}.{today1.split('.')[2][-2:]}"
                weekends = [today]

                for i in range(cur_day_index - 1, 0, -1):
                    if calendar['Type'].iloc[i] == 'Working':
                        yesterday1 = calendar['Day'].iloc[i]
                        break

                    weekends.append(calendar['Day'].iloc[i][:6] + '20' + calendar['Day'].iloc[i][-2:])

                # logger.info(weekends)
                # logger.info(weekends)

                logger.info('Start day', today)
                logger.info(weekends)
                for day in weekends[::-1]:

                    logger.info('Start day', today)
                    logger.info(weekends)

                    # logger.warning(f'Started processing {day}')
                    today = datetime.datetime.now().date()

                    day_ = int(day.split('.')[0])
                    month_ = int(day.split('.')[1])
                    year_ = int(day.split('.')[2])

                    today = datetime.date(year_, month_, day_)

                    cashbook_day = (today - datetime.timedelta(days=5)).strftime('%d.%m.%Y')

                    days_ = []

                    for i in range(7, 1, -1):
                        day = (today - datetime.timedelta(days=i)).strftime('%d.%m.%Y')
                        days_.append(day)

                    today = today.strftime('%d.%m.%Y')

                    # print(f'Cahsbook: {cashbook_day}')
                    # logger.warning(f'Cahsbook: {cashbook_day}')
                    logger.warning(days_)
                    logger.info('==========================\n')

                    logger.info(days_)

                    net_use(Path(template_path).parent, owa_username, owa_password)
                    net_use(ecp_paths, owa_username, owa_password)

                    # tg_send(f'Робот запустился - <b>{today}</b>\n\nДата для выгрузки чеков из Спрута - <b>{cashbook_day}</b>\n\nДата проверки в 1С - <b>{days}</b>', bot_token=tg_token, chat_id=chat_id)

                    try:

                        logger.warning(f'Начали отработку за {today} | {today1}')
                        # * ----- 1 -----
                        logger.warning('Начали Спрут')
                        logger.info('Начали Спрут')

                        logger.warning(f'Cashbook: {cashbook_day}')

                        filepath, bonuses = open_cashbook(cashbook_day)

                        if filepath == '':
                            smtp_send(fr"""Добрый день!
                            Сверка ОПТа за {today} завершилась - Пусто в Розничных чеках""",
                                      to=email_to,
                                      subject=f'Сверка ОПТа за {today}', username=smtp_author, url=smtp_host)
                            logger.warning(f'Законичили отработку за {today} - Пусто в Розничных чеках')
                        #     continue

                        # # ! Uncomment, if you are compiling for the virtual machines
                        filepath = filepath.replace('Documents', 'Downloads')

                        # main_file = r'\\vault.magnum.local\Common\Stuff\_06_Бухгалтерия\Для робота\Процесс Сверка ОПТа\Файл сбора Апрель 2024.xlsx'
                        # bonuses = []
                        # filepath = r'C:\Users\Abdykarim.D\Downloads\magnumopt_2023-09-02_2023-09-03.xlsx'
                        # * ----- 2 -----
                        main_file = create_collection_file(filepath, today, bonuses)
                        Path(filepath).unlink()

                        # # * ----- 3 -----
                        logger.warning('Начали Epay')
                        logger.info('Начали Epay')
                        filepath = None
                        for tries in range(5):
                            with suppress(Exception):
                                filepath = homebank(homebank_login, homebank_password, days_[0], days_[-1])

                                logger.warning('Закончили Epay')
                                logger.info('Зкончили Epay')
                                break

                        logger.warning(filepath)
                        logger.info(filepath)
                        check_homebank_and_collection(filepath, main_file)
                        Path(filepath).unlink()

                        # * ----- 4 -----
                        logger.warning('Начали 1C')
                        logger.info('Начали 1C')

                        for tries in range(5):
                            try:
                                all_days = odines_part(days_)

                                odines_check_with_collection(all_days, days_, main_file)
                                break

                            except Exception as err:
                                print("ERROR 1C:", err)
                                logger.warning(f"ERROR OCCURED 1C: {err}")
                                traceback.print_exc()

                        # * ----- 5 -----
                        logger.warning('Начали ОФД')
                        logger.info('Начали ОФД')

                        for tries in range(5):
                            try:
                                ofd_distributor(main_file)
                                break
                            except Exception as err1:
                                logger.warning(f"ERROR OCCURED HERE OFD: {err1}")
                                traceback.print_exc()

                        logger.warning(f'Законичили отработку за {today}')
                        logger.info(f'Законичили отработку за {today}')

                        smtp_send(fr"""Добрый день!
                                       Сверка ОПТа за {today} завершилась успешно, файл сбора лежит в папке {main_file}""",
                                  to=email_to,
                                  subject=f'Сверка ОПТа за {today}', username=smtp_author, url=smtp_host)
                        # , 'Sagimbayeva@magnum.kz', 'Ashirbayeva@magnum.kz'

                    except Exception as err:
                        traceback.print_exc()
                        logger.info(f'ERRROR!!!, {err}')
                        logger.warning(f'ERRROR!!!, {err}')
                else:
                    print(1)

    except Exception as main_error:
        traceback.print_exc()
        print(f'MAIN ERROR OCCURED: {main_error} ||| {traceback.format_exc()}')
        logger.info(f'MAIN ERROR OCCURED: {main_error}')
        logger.warning(f'MAIN ERROR OCCURED: {main_error}')
