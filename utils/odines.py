from contextlib import suppress
from time import sleep

from openpyxl import load_workbook

from config import logger, ip_address
from core import Odines
from tools.clipboard import clipboard_set, clipboard_get
from utils.check_time_diff import check_time_diff


def odines_part(days):

    opened_table_selector = {"title": "", "class_name": "", "control_type": "Table", "visible_only": True,
                             "enabled_only": True, "found_index": 0}
    filter_selector = {"title": "Установить отбор и сортировку списка...", "class_name": "",
                       "control_type": "Button",
                       "visible_only": True, "enabled_only": True, "found_index": 0}
    filter_whole_wnd_selector = {"title": "Отбор и сортировка", "class_name": "V8NewLocalFrameBaseWnd",
                                 "control_type": "Window", "visible_only": True, "enabled_only": True,
                                 "found_index": 0}

    path = r'C:\Program Files\1cv8\8.3.13.1644\bin\1cv8.exe' if ip_address == '172.20.1.24' else r'C:\Program Files\1cv8\common\1cestart.exe'
    app = Odines(path)
    app.run()

    sleep(5)
    app.maximize_inner()

    app.navigate("Банк и касса", "Отчет банка по операциям эквайринга", maximize_innder=True)

    table_element = app.find_element(opened_table_selector)

    app.find_element(filter_selector).click()

    app.parent_switch(filter_whole_wnd_selector, resize=True)

    sleep(1)

    app.find_element({"title": "Пометка удаления", "class_name": "", "control_type": "CheckBox",
                      "visible_only": True, "enabled_only": True, "found_index": 0}).click()
    app.find_element({"title": "", "class_name": "", "control_type": "Edit",
                      "visible_only": True, "enabled_only": True, "found_index": 3}).type_keys('Нет', app.keys.TAB,
                                                                                               protect_first=True, clear=True,
                                                                                               click=True)

    app.find_element({"title": "Организация", "class_name": "", "control_type": "CheckBox",
                      "visible_only": True, "enabled_only": True, "found_index": 0}).click()

    app.find_element({"title": "", "class_name": "", "control_type": "Edit",
                      "visible_only": True, "enabled_only": True, "found_index": 7}).type_keys('ТОО "Magnum Cash&Carry"', app.keys.TAB,
                                                                                               protect_first=True, clear=True,
                                                                                               click=True)

    app.find_element({"title": "Контрагент", "class_name": "", "control_type": "CheckBox",
                      "visible_only": True, "enabled_only": True, "found_index": 0}).click()

    app.find_element({"title": "", "class_name": "", "control_type": "Edit",
                      "visible_only": True, "enabled_only": True, "found_index": 37}).type_keys('Частное лицо- ОПТ', app.keys.TAB,
                                                                                                protect_first=True, clear=True,
                                                                                                click=True)

    app.find_element({"title": "OK", "class_name": "", "control_type": "Button",
                      "visible_only": True, "enabled_only": True, "found_index": 0}).click()

    app.parent_back(1)

    app.find_element({"title": "", "class_name": "", "control_type": "Table", "visible_only": True, "enabled_only": True, "found_index": 0}).click()

    # for _ in range(15):
    #     app.find_element({"title": "", "class_name": "", "control_type": "Table", "visible_only": True, "enabled_only": True, "found_index": 0}).type_keys(app.keys.PAGE_UP)

    found = False
    # for _ in range(6):

    els = app.find_elements({"title_re": ".* Дата", "class_name": "", "control_type": "Custom",
                             "visible_only": True, "enabled_only": True}, timeout=3)

    all_days = []

    for i in els:

        clipboard_set("")
        i.type_keys("^c", click=True, clear=False)

        get_report_date = clipboard_get()
        get_report_date = str(get_report_date).strip()[:10]
        logger.info(get_report_date)

        if get_report_date in days:

            found = True

            transaction_dict = dict()

            i.click(double=True)

            sleep(0.5)

            app.parent_switch({"title": "", "class_name": "", "control_type": "Pane",
                               "visible_only": True, "enabled_only": True}, resize=True, set_focus=True, maximize=True)

            with suppress(Exception):
                app.find_element({"title": "Развернуть", "class_name": "", "control_type": "Button",
                                  "visible_only": True, "enabled_only": True}, timeout=3).click()

            # ? Собираем все даты транзакций и их суммы
            transactions = app.find_elements({"title_re": ".* Дата транзакции$", "class_name": "", "control_type": "Custom",
                                              "visible_only": True, "enabled_only": True}, timeout=10)

            print(transactions)

            print(len(transactions))

            summs = app.find_elements({"title_re": ".* Сумма$", "class_name": "", "control_type": "Custom",
                                       "visible_only": True, "enabled_only": True}, timeout=5)

            print(summs)
            print(len(summs))

            for ind, transaction in enumerate(transactions):
                if True:
                    logger.info('-------------------------------------------')
                    clipboard_set("")
                    transaction.type_keys("^c", click=True, clear=False)
                    transaction.type_keys(app.keys.DOWN, click=True, clear=False)

                    transaction_date = clipboard_get()
                    transaction_date = str(transaction_date).strip()
                    logger.info(f'Transaction {transaction}: {transaction_date}')

                    clipboard_set("")
                    logger.info('Clicking on', ind, summs[ind])
                    summs[ind].type_keys("^c", click=True, clear=False)

                    summ = clipboard_get()
                    summ = round(float(str(summ).replace(' ', '').replace(',', '.').replace(' ', '')))
                    logger.info('Sum:', summ)
                    logger.info('-------------------------------------------')

                    transaction_dict.update({transaction_date: summ})

                    if ind == 17:
                        for scroll in range(10):
                            transaction.type_keys(app.keys.DOWN, click=True, clear=False)
                # except Exception as err:
                #     print(f'ERROR: {err}')

            app.find_element({"title": "Закрыть", "class_name": "", "control_type": "Button",
                              "visible_only": True, "enabled_only": True}).click()

            app.parent_back(1)

            all_days.append(transaction_dict)

        elif found:
            break

        # app.find_element({"title": "", "class_name": "", "control_type": "Table", "visible_only": True, "enabled_only": True, "found_index": 0}).click()
        #
        app.find_element({"title": "", "class_name": "", "control_type": "Table", "visible_only": True, "enabled_only": True, "found_index": 0}).type_keys(app.keys.DOWN)
        # app.find_element({"title": "", "class_name": "", "control_type": "Table", "visible_only": True, "enabled_only": True, "found_index": 0}).type_keys(app.keys.PAGE_DOWN)

    app.quit()
    logger.warning('FINISHING 1C')
    logger.info(all_days)
    logger.warning('FINISHING 1C.1')
    return all_days


def odines_check_with_collection(all_days_, main_file):
    logger.warning('Started checking 1C')
    collection_file = load_workbook(main_file)

    collection_sheet = collection_file['Файл сбора']

    logger.info(collection_sheet.max_row)

    for row in range(3, collection_sheet.max_row + 1):

        if collection_sheet[f'F{row}'].value is not None:
            continue

        collection_sheet[f'F{row}'].value = 'нет'
        for day_ in all_days_:
            for single_day in day_:

                print(collection_sheet[f'C{row}'].value)
                time_diff = check_time_diff(collection_sheet[f'C{row}'].value, single_day, 5)

                if time_diff and abs(day_.get(single_day) - round(collection_sheet[f'D{row}'].value)) <= 1:
                    logger.info('--------------------------------------------------------------------------')
                    logger.info(f"{single_day}, {collection_sheet[f'C{row}'].value}, {day_.get(single_day)},"
                                f"{collection_sheet[f'D{row}'].value}, {time_diff}")
                    collection_sheet[f'F{row}'].value = 'да'

    collection_file.save(main_file)
    logger.info('--------------------------------------------------------------------------')
