import os
from time import sleep

from openpyxl import load_workbook
import pandas as pd

from config import logger, download_path, months
from utils.check_time_diff import check_time_diff
from tools.web import Web


def homebank(email, password, start_date, end_date):

    web = Web()
    web.run()
    web.get('https://epay.homebank.kz/login')

    for tries in range(3):
        if web.wait_element('//*[@id="mp-content"]/section/main/div[2]/div/div/div[2]/form/div[1]/div/div/span/div/input'):
            break
        else:
            web.driver.refresh()

    sleep(5)

    web.find_element('//*[@id="mp-content"]/section/main/div[2]/div/div/div[2]/form/div[1]/div/div/span/div/input').type_keys(email)
    web.find_element('//*[@id="mp-content"]/section/main/div[2]/div/div/div[2]/form/div[2]/div/div/span/div/span/input').type_keys(password)

    web.find_element('//*[@id="mp-content"]/section/main/div[2]/div/div/div[2]/form/div[3]/div/div/span/button').click()

    web.wait_element("//span[@class='src-layouts-main-header_button hint-section-1-step-3']")

    web.get('https://epay.homebank.kz/statements/payment')

    web.find_element("//span[contains(text(), '427693/14-EC27/07')]").click()

    web.find_element('//*[@id="mp-content"]/div/div/div/div/div[1]/div/div/div[1]/div/div/div/div[2]/button').click()
    sleep(1)
    web.find_element('//*[@id="period"]').click()

    sleep(1)

    day_ = int(start_date.split('.')[0])
    month_ = int(start_date.split('.')[1])
    year_ = start_date.split('.')[2]

    start_ = f"{day_} {months[month_ - 1]} {year_} г."

    day_ = int(end_date.split('.')[0])
    month_ = int(end_date.split('.')[1])
    year_ = end_date.split('.')[2]

    end_ = f"{day_} {months[month_ - 1]} {year_} г."

    logger.info(f"//td[@title = '{start_}']")
    logger.info(f"//td[@title = '{end_}']")

    # ? Нажимает на нужные даты в календаре
    for tries in range(15):
        try:
            web.find_element(f"//td[@title = '{start_}']", timeout=5).click()
            web.find_element(f"//td[@title = '{end_}']", timeout=2).click()
            break

        except:
            web.find_element("//a[contains(@title, 'Предыдущий месяц')]").click()

    web.execute_script_click_xpath_selector("//span[contains(text(), 'XLSX')]")

    web.execute_script_click_xpath_selector("//button[contains(@class, 'ant-btn ant-btn-primary ant-btn-lg')]")  # Form the report
    # Нижняя стrрока - кнопка Отменить, использовалось в тесте, чтобы не формировать один и тот же отчёт по несколько раз
    # web.execute_script_click_xpath_selector("//button[contains(@class, 'ant-btn ant-btn-lg')]") # ant-btn ant-btn-primary ant-btn-lg

    logger.info('started waiting')
    sleep(25)

    web.find_element("(//span[@class='src-pages-statements-styles_status-column'])[1]").click()
    logger.info('clicked downloading')
    filepath = None
    found = False
    for _ in range(150):
        for file in os.listdir(download_path):
            if 'magnumopt' in file and '$' not in file and '.crdownload' not in file:
                filepath = os.path.join(download_path, file)
                found = True
                break
        if found:
            break

        sleep(1)

    return filepath


def check_homebank_and_collection(filepath_, main_file):

    collection_file = load_workbook(main_file)

    collection_sheet = collection_file['Файл сбора']

    count = 0
    for row in range(3, collection_sheet.max_row + 1):

        if collection_sheet[f'E{row}'].value is not None:
            continue

        try:
            df = pd.read_excel(filepath_)

            df.columns = df.iloc[10]
        except:
            collection_sheet[f'E{row}'].value = 'нет'
            continue

        try:
            new_df = df[df['Дата валютир.'] == collection_sheet[f'B{row}'].value.strftime("%d.%m.%Y")]
        except:
            new_df = df[df['Дата валютир.'] == collection_sheet[f'B{row}'].value]

        filtered_df = new_df[new_df['Оплачено'] == collection_sheet[f'D{row}'].value]  # Отобрал только те записи, которые были произведены за D{row} день из файла сбора

        collection_sheet[f'E{row}'].value = 'нет'
        # logger.info(filtered_df)
        print("DF:")
        print(new_df['Дата/время транз.'])
        print(filtered_df)
        print()
        for ind, times in enumerate(new_df['Дата/время транз.']):
            print('###', ind, times)
            collection_date, homebank_date = collection_sheet[f'C{row}'].value, times

            time_diff = check_time_diff(collection_date, homebank_date, 5)
            print('TIME', time_diff, (collection_sheet[f'D{row}'].value == new_df['Оплачено'].iloc[ind]), (collection_sheet[f'D{row}'].value - int(collection_sheet[f'I{row}'].value) == new_df['Оплачено'].iloc[ind]))
            if time_diff and (collection_sheet[f'D{row}'].value == new_df['Оплачено'].iloc[ind] or collection_sheet[f'D{row}'].value - int(collection_sheet[f'I{row}'].value) == new_df['Оплачено'].iloc[ind]):
                logger.info(time_diff)
                collection_sheet[f'E{row}'].value = 'да'
                break
        count += 1

    collection_file.save(main_file)
