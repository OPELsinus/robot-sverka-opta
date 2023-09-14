import datetime
import os
from contextlib import suppress
from time import sleep

import pandas as pd
from openpyxl import load_workbook

from pywinauto import keyboard

from config import logger, ecp_paths, mapping_path
from tools.app import App
from tools.web import Web
from utils.check_time_diff import check_time_diff


def sign_ecp_kt(ecp):

    app = App('')

    el = {"title": "Открыть файл", "class_name": "SunAwtDialog", "control_type": "Window",
          "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}

    if app.wait_element(el, timeout=30):

        keyboard.send_keys(ecp.replace('(', '{(}').replace(')', '{)}'), pause=0.01, with_spaces=True)
        sleep(0.05)
        keyboard.send_keys('{ENTER}')

        logger.info('Finished ECP')

        app = None

        return 'signed'

    else:
        logger.info('Quit mazafaka')
        app = None
        return 'broke'


def sign_ecp_trans(ecp):

    app = App('')

    el = {"title": "Открыть файл", "class_name": "SunAwtDialog", "control_type": "Window",
          "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}

    if app.wait_element(el, timeout=30):

        keyboard.send_keys(ecp.replace('(', '{(}').replace(')', '{)}'), pause=0.01, with_spaces=True)
        sleep(0.05)
        keyboard.send_keys('{ENTER}')

        if app.wait_element({"title_re": "Формирование ЭЦП.*", "class_name": "SunAwtDialog", "control_type": "Window",
                             "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}, timeout=30):
            app.find_element({"title_re": "Формирование ЭЦП.*", "class_name": "SunAwtDialog", "control_type": "Window",
                              "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}).type_keys('Aa123456')
            sleep(2)
            keyboard.send_keys('{ENTER}')
            sleep(3)
            keyboard.send_keys('{ENTER}')
            app = None
            return 'signed'
        else:
            logger.info('Quit mazafaka1')
            app = None
            return 'broke'
    else:
        logger.info('Quit mazafaka')
        app = None
        return 'broke'


def ofd_distributor(main_file):

    collection_file = load_workbook(main_file)

    collection_sheet = collection_file['Файл сбора']

    mapping_file = pd.read_excel(mapping_path)
    logger.info(mapping_file.columns)
    for row in range(3, collection_sheet.max_row + 1):

        print(collection_sheet[f'G{row}'].value)

        if collection_sheet[f'G{row}'].value is not None:
            continue

        seacrh_date = collection_sheet[f'B{row}'].value
        collection_sheet[f'G{row}'].value = 'нет'
        print(f"{row} | {collection_sheet[f'A{row}'].value}")
        short_name = mapping_file[mapping_file['Наименование в Спруте'].str.lower() == collection_sheet[f'A{row}'].value.lower()]['Филиал'].iloc[0]
        ecp_path = mapping_file[mapping_file['Наименование в Спруте'].str.lower() == collection_sheet[f'A{row}'].value.lower()]['Площадка в Спруте'].iloc[0]

        ecp_auth = ''
        ecp_sign = ''

        for file in os.listdir(os.path.join(ecp_paths, ecp_path)):
            if 'AUTH' in str(file):
                ecp_auth = os.path.join(ecp_paths, ecp_path, file)
            if 'GOST' in str(file):
                ecp_sign = os.path.join(ecp_paths, ecp_path, file)
        # logger.info(ecp_sign, ecp_auth)
        print(short_name)

        # * Get the mapping which operator to use for that branch
        ofd_operator = mapping_file[mapping_file['Наименование в Спруте'].str.lower() == collection_sheet[f'A{row}'].value.lower()]['ОФД'].iloc[0]

        if ofd_operator == 'Казахтелеком':

            print('kazakhtelekom')
            collection_sheet[f'G{row}'].value = 'нет'
            try:
                open_oofd_kazakhtelekom(seacrh_date, collection_sheet, row, ecp_auth, ecp_sign)
            except:
                logger.warning(f'FAILED OFD at {seacrh_date}, {row}, {ecp_auth}')
                raise Exception('FAILED OFD')

        elif ofd_operator == 'Транстелеком':

            print('trans')
            collection_sheet[f'G{row}'].value = 'нет'
            try:
                open_oofd_trans(seacrh_date, collection_sheet, row, ecp_auth, ecp_sign)
            except:
                logger.warning(f'FAILED OFD at {seacrh_date}, {row}, {ecp_auth}')
                raise Exception('FAILED OFD')

    collection_file.save(main_file)


def open_oofd_trans(seacrh_date, collection_sheet, row, ecp_auth, ecp_sign):

    web = Web()

    web.run()
    web.get('https://ofd1.kz/login')

    web.find_element('//*[@id="login_by_cert_btn"]').click()
    sign_ecp_trans(ecp_auth)

    # if web.wait_element('//*[@id="close_i_modal"]/img', timeout=10):
    #     web.find_element('//*[@id="close_i_modal"]/img').click()
    sleep(5)
    web.get('https://ofd1.kz/cash_register?status_type=registered')

    web.find_element('//*[@id="sample"]').type_keys(collection_sheet[f'K{row}'].value)

    web.find_element('//*[@id="sign_btn"]').click()

    web.find_element('//*[@id="serach_results"]//a').click()

    web.find_element('//*[@id="shift_list_button"]').click()

    day_ = int(seacrh_date.split('.')[0])
    month_ = int(seacrh_date.split('.')[1])
    year_ = int(seacrh_date.split('.')[2])

    seacrh_date = datetime.datetime(year_, month_, day_)

    # ? Изменяем значения дат на нужное нам
    web.set_elements_value(xpath='//*[@id="start_date"]', value=seacrh_date.strftime('%Y-%m-%d'))
    web.set_elements_value(xpath='//*[@id="end_date"]', value=(seacrh_date + datetime.timedelta(days=1)).strftime('%Y-%m-%d'))
    logger.info(seacrh_date)
    # web.find_element('//*[@id="shift_list_button_list"]').click()
    web.execute_script_click_xpath_selector('//*[@id="shift_list_button_list"]')

    sleep(1.5)

    if web.wait_element('//*[@id="shifts-container"]/tr/td[5]', timeout=5):

        dates = web.find_elements('//*[@id="shifts-container"]/tr/td[5]/preceding-sibling::td[3]')
        summs = web.find_elements('//*[@id="shifts-container"]/tr/td[5]')

        for ind in range(len(dates)):
            summ_ = round(float(summs[ind].get_attr('text').replace(' ', '')))
            # Сори за такие длинные выражения xD
            time_diff = check_time_diff(collection_sheet[f'C{row}'].value, datetime.datetime.strptime(dates[ind].get_attr('text'), '%Y-%m-%d %H:%M:%S').strftime('%d.%m.%Y %H:%M:%S'), 5)

            if summ_ == int(collection_sheet[f'D{row}'].value) and time_diff:
                logger.info(dates[ind].get_attr('text'))
                logger.info(summs[ind].get_attr('text'))
                collection_sheet[f'G{row}'].value = 'да'
                logger.info('----------------------------------------------')
        sleep(1)


def open_oofd_kazakhtelekom(seacrh_date, collection_sheet, row, ecp_auth, ecp_sign):

    web = Web()

    web.run()
    web.get('https://org.oofd.kz/#/landing/eds-login')

    if web.wait_element("//button[contains(text(), 'kz')]", timeout=10):
        web.find_element("//button[contains(text(), 'kz')]").click()
        web.execute_script_click_xpath_selector("//div[contains(text(), 'RU')]")

    web.find_element("//button[contains(text(), 'Войти с ЭЦП')]").click()

    web.find_element('//*[@id="storage-password"]').type_keys('Aa123456')
    web.execute_script_click_xpath_selector('//*[@id="storage-type"]/div/div[2]/div/p[2]/span')
    sleep(2)

    sign_ecp_kt(ecp_auth)

    sleep(3)
    if web.wait_element("//button[contains(text(), 'Проверить')]", timeout=5):
        web.find_element("//button[contains(text(), 'Проверить')]").click()

    with suppress(Exception):
        app = App('')

        app.find_element({"title": "Ввод пароля", "class_name": "SunAwtDialog", "control_type": "Window",
                          "visible_only": True, "enabled_only": True, "found_index": 0}, timeout=15).type_keys('Aa123456', app.keys.ENTER)

        sleep(1)

        app.find_element({"title": "Ввод пароля", "class_name": "SunAwtDialog", "control_type": "Window",
                          "visible_only": True, "enabled_only": True, "found_index": 0}, timeout=10).type_keys('Aa123456', app.keys.ENTER)
        app = None

    web.find_element("//input[contains(@placeholder, 'Магазин, касса')]").type_keys(str(collection_sheet[f'K{row}'].value).replace(' ', ''), web.keys.ENTER)  # Filling serial number

    web.find_element("(//a[@class='kkm'])[1]").click()  # Find & click on the first element
    # sleep(100)
    logger.info(f"{seacrh_date} {seacrh_date.split('.')}")

    year = seacrh_date.split('.')[2]
    month = seacrh_date.split('.')[1]
    day = seacrh_date.split('.')[0]

    # ? С помощью JS меняем поля дат и url на нужную нам дату
    web.set_elements_innerhtml_or_value('//*[@id="mat-input-0"]', element_type='value', date=f'{year}-{month}-{day}T00:00:00', value=f'{int(day)}.{int(month)}.{year}')
    web.set_elements_innerhtml_or_value('//*[@id="mat-input-1"]', element_type='value', date=f'{year}-{month}-{day}T23:59:59', value=f'{int(day)}.{int(month)}.{year}')
    web.set_elements_innerhtml_or_value("//input[@ng-reflect-name='shiftNumber']", element_type='innerHTML', value='')
    print('URL:', web.driver.current_url, end=' - ')
    new_url = web.driver.current_url.split('?')[0] + f'?startDate={year}-{month}-{day}T00:00:00&endDate={year}-{month}-{day}T23:59:59&page=1'
    print(new_url)
    web.get(new_url)
    web.get(new_url)
    web.driver.refresh() # Обновляем, чтобы данные точно прогрузились

    transactions = web.find_elements("//div[@class='transaction-wrapper ng-star-inserted']", timeout=40)
    times = web.find_elements("//div[@class='transaction-wrapper ng-star-inserted']/tax-transaction/div/span/span", timeout=1)
    summs = web.find_elements("//div[@class='transaction-wrapper ng-star-inserted']//span[@class='transaction__sum ng-star-inserted']", timeout=1)

    for ind in range(len(transactions)):

        time_ = " " + times[ind].get_attr('text').split()[-1] + ":00" # Дописываем :00 в конец, чтобы преобразовать в формату datetime для сравнения
        summ_ = round(float(summs[ind].get_attr('text').replace('₸', '').replace(' ', '').replace(',', '.')))

        sleep(.1)

        if check_time_diff(seacrh_date + time_, collection_sheet[f'C{row}'].value, 5) and summ_ == int(collection_sheet[f'D{row}'].value):
            logger.info(f"{seacrh_date + time_} {summ_}")
            logger.info(f"{check_time_diff(seacrh_date + time_, collection_sheet[f'C{row}'].value, 5)}")
            collection_sheet[f'G{row}'].value = 'да'

    # sleep(10000)
    logger.info('-----------------------------------------------------------------------------')
