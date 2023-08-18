import datetime
import socket
import sys

import warnings
from contextlib import suppress
from pathlib import Path
from shutil import move
from time import sleep

import openpyxl
import pandas as pd

from logs import init_logger

warnings.filterwarnings('ignore')
from openpyxl import load_workbook
import psycopg2
from config import monthes, sTempPath, db_user, db_pass, db_host, db_port, db_name, template_path, sprut_base, \
    razbor_save_path, global_path, smtp_author, to_whom, smtp_host, \
    system_info, tg_token
from core import Odines, Sprut
from rpamini import BusinessException, check_file_downloaded, fix_excel_file_error, get_hostname, \
    clipboard_get, \
    clipboard_set, try_except_decorator, protect_path, send_telegram, send_message_by_smtp, retry_n_times
import time
import traceback


@try_except_decorator
def process(row):
    logger.info("----------------------------------------------start process--------------------------------")
    try:
        logger.info(row)
        start_time = time.time()

        date_string = row[1]

        date_format = "%d.%m.%Y"
        date_object = datetime.datetime.strptime(date_string, date_format)
        tr: Transaction = Transaction(id=row[0], process_date=date_object, branch_name=row[2], odines_name=row[3],
                                      sprut_name=row[4], store_names=row[5], main_excel_file=row[6], status=row[7],
                                      retry_count=row[8])
        print(tr.__dict__)
        tr.status = "Retried"
        tr.retry_count = int(tr.retry_count) + 1
        report_path = sTempPath.joinpath(f"100912 Сверка_Z_отчётов_и_оборота_Спрут {protect_path(tr.branch_name)}.xlsx")
        tr.get_sprut(report_path=report_path)
        z_sum, sprut_sum = tr.get_values_from_912(report_path)
        if "REF" in str(z_sum) or "REF" in str(sprut_sum):
            logger.info(f"100912 Сверка_Z_отчётов_и_оборота_Спрут {tr.branch_name}  Business Exception")
            tr.comments = "Пуст 100912 Спрут"
            tr.status = "Success"
            tr.update_set_status()
            return True
        app = Odines()
        app.run()
        tr.app = app
        tr.download_all_reports()
        print("Exiting 1C")
        app.quit()

        if "КЗФ" in tr.branch_name:
            logger.info("Это КЗО")
            total_sum, razbor_path = tr.read_downloaded_excel_files_kzo()
        else:
            total_sum, razbor_path = tr.read_downloaded_excel_files()

        comment = f"total sum - sprut_sum: {float(total_sum) - float(sprut_sum)}"
        logger.info(f"total sum: {total_sum}, sprut_sum: {sprut_sum}, z_beznal: {z_sum}")
        print(f"************{comment}******************")

        if total_sum == sprut_sum:
            logger.info("Нет расхождении")
            tr.status = 'Success'
            tr.comments = f"Нет расхождении {comment}"
            tr.update_set_status()
            end_time = time.time()
            tr.edit_main_excel_file(z_sum, sprut_sum, total_sum)
            logger.info(f"<---------------Execution Time: {end_time - start_time:.4f} seconds----------->")
            return True

        logger.info("Есть расхождения")
        tr.comments = f" Расхождения: {comment}"
        logger.info("Запуск Спрута для отчета 101160")

        report_path = sTempPath.joinpath(
            f"Выручка по кассам с выбором форм оплат 101160 {protect_path(tr.branch_name)}.xlsx")

        tr.get_sprut_101160(report_path=report_path)

        tr.data_manipulation_101160(report_path, razbor_path)

        tr.edit_main_excel_file(z_sum, sprut_sum, total_sum)

        tr.status = 'Success'
        tr.update_set_status()
        end_time = time.time()
        logger.info(f"<---------------Execution Time: {end_time - start_time:.4f} seconds----------->")
        logger.info('-----------------------------End of process-----------------------------')
    except Exception as ex:
        traceback.print_exc()
        tb = traceback.extract_tb(ex.__traceback__)
        filename, line, func, text = tb[-1]
        msg = f'Error: {func} {ex}'
        logger.error(msg)
        tr.status = "Retried"
        tr.error_message = msg
        tr.update_set_status()
        raise ex


class Transaction:
    def __init__(self, id, process_date, branch_name, odines_name, sprut_name, store_names, main_excel_file, status,
                 retry_count):
        self.id = id
        self.process_date = process_date
        self.branch_name = branch_name
        self.odines_name = odines_name
        self.sprut_name = sprut_name
        self.store_names = store_names
        self.main_excel_file = main_excel_file
        self.errors = None
        self.comments = None
        self.status = status
        self.app = None
        self.retry_count = retry_count
        self.reports = []

    def download_all_reports(self):
        logger.info('starting 1c')
        self.app.navigate("Банк и касса", "Отчет банка по операциям эквайринга", maximize_innder=True)

        str_date_from = (self.process_date - datetime.timedelta(days=2)).strftime("%d.%m.%Y")
        str_date_to = (self.process_date + datetime.timedelta(days=2)).strftime("%d.%m.%Y")
        self.app.find_element(
            {"title": "Установить интервал дат...", "class_name": "", "control_type": "Button", "visible_only": True,
             "enabled_only": True, "found_index": 0}).click()
        self.app.parent_switch(
            {"title": "Настройка периода", "class_name": "V8NewLocalFrameBaseWnd", "control_type": "Window",
             "visible_only": True, "enabled_only": True, "found_index": 0})
        self.app.find_element(
            {"title": "", "class_name": "", "control_type": "RadioButton", "visible_only": True, "enabled_only": True,
             "found_index": 0}).click()

        self.app.find_element(
            {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
             "found_index": 0}).type_keys(str_date_from, self.app.keys.TAB, click=True, clear=True, set_focus=False)
        self.app.find_element(
            {"title": "", "class_name": "", "control_type": "RadioButton", "visible_only": True, "enabled_only": True,
             "found_index": 1}).click()

        self.app.find_element(
            {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
             "found_index": 1}).type_keys(str_date_to, self.app.keys.TAB, click=True, clear=True, set_focus=False)
        self.app.find_element(
            {"title": "OK", "class_name": "", "control_type": "Button", "visible_only": True, "enabled_only": True,
             "found_index": 0}).click()
        # * End ----------------------------------------------------------------------------------------------------------------------------------------------------------------------------

        self.app.parent_back(1)

        self.app.find_element(
            {"title": "Установить отбор и сортировку списка...", "class_name": "", "control_type": "Button",
             "visible_only": True, "enabled_only": True, "found_index": 0}).click()
        self.app.parent_switch(
            {"title": "Отбор и сортировка", "class_name": "V8NewLocalFrameBaseWnd", "control_type": "Window",
             "visible_only": True, "enabled_only": True, "found_index": 0}, maximize=True)

        self.app.find_element(
            {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
             "found_index": 3}).type_keys("Нет", self.app.keys.TAB, click=True, clear=True)
        self.app.find_element(
            {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
             "found_index": 7}).type_keys(self.odines_name, self.app.keys.TAB * 25, click=True, clear=True,
                                          protect_first=True)
        self.app.find_element(
            {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
             "found_index": 35}).type_keys("Нет", self.app.keys.TAB, click=True, clear=True)
        self.app.find_element(
            {"title": "OK", "class_name": "", "control_type": "Button", "visible_only": True, "enabled_only": True,
             "found_index": 0}).click()
        self.app.parent_back(1)

        trans_exists = self.app.wait_element(
            {"title_re": ".* Номер", "class_name": "", "control_type": "Custom", "visible_only": True,
             "enabled_only": True, "found_index": 0}, timeout=3)
        if not trans_exists:
            logger.info("Нет транзакции")
            raise BusinessException("Нет транзакции", "download_all_reports")
        n_trans: int = 0
        while trans_exists:
            clipboard_set("")
            self.app.find_element(
                {"title_re": ".* Номер", "class_name": "", "control_type": "Custom", "visible_only": True,
                 "enabled_only": True, "found_index": n_trans}).type_keys("^c", click=True, clear=False)
            doc_num = clipboard_get()

            clipboard_set("")
            self.app.find_element(
                {"title_re": ".* Дата", "class_name": "", "control_type": "Custom", "visible_only": True,
                 "enabled_only": True, "found_index": n_trans}).type_keys("^c", click=True, clear=False)
            get_report_date = clipboard_get()
            get_report_date = str(get_report_date).strip()[:10]

            clipboard_set("")
            self.app.find_element(
                {"title_re": ".* Эквайер", "class_name": "", "control_type": "Custom", "visible_only": True,
                 "enabled_only": True, "found_index": n_trans}).type_keys("^c", click=True, clear=False)
            get_bank = clipboard_get()
            if "народный" in get_bank.lower():
                get_bank = "народный"
            elif 'kaspi' in get_bank.lower():
                get_bank = "kaspi"
            else:
                get_bank = "другой банк"

            clipboard_set("")
            self.app.find_element(
                {"title_re": ".* Сумма возвратов", "class_name": "", "control_type": "Custom", "visible_only": True,
                 "enabled_only": True, "found_index": n_trans}).type_keys("^c", click=True, clear=False)
            returns_exist = clipboard_get()
            # * type path where the table is stored as xlsx file

            tmp_file_path = sTempPath.joinpath(
                f"Поступления оплат от клиентов {protect_path(self.branch_name)} {get_report_date} {doc_num} {get_bank}.xlsx")

            # ^ check where we already downloaded it
            if tmp_file_path not in self.reports:

                self.app.find_element(
                    {"title_re": ".* Номер", "class_name": "", "control_type": "Custom", "visible_only": True,
                     "enabled_only": True, "found_index": n_trans}).click(double=True)

                self.app.parent_switch(
                    {"title": "Поступления оплат от клиентов", "class_name": "", "control_type": "Tab",
                     "visible_only": True, "enabled_only": True, "found_index": 0})
                self.app.find_element(
                    {"title_re": ".* Дата платежа", "class_name": "", "control_type": "Custom", "visible_only": True,
                     "enabled_only": True, "found_index": 0}).click(right=True)
                self.app.parent_back(1)
                self.app.find_element(
                    {"title": "Вывести список...", "class_name": "", "control_type": "MenuItem", "visible_only": True,
                     "enabled_only": True, "found_index": 0}).click()
                self.app.parent_switch(
                    {"title": "Вывести список", "class_name": "V8NewLocalFrameBaseWnd", "control_type": "Window",
                     "visible_only": True, "enabled_only": True, "found_index": 0})
                self.app.find_element(
                    {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
                     "found_index": 0}).type_keys(self.app.keys.DOWN, click=True, clear=False)
                self.app.find_element(
                    {"title": "Табличный документ", "class_name": "", "control_type": "ListItem", "visible_only": True,
                     "enabled_only": True, "found_index": 0}).click()
                self.app.find_element({"title": "ОК", "class_name": "", "control_type": "Button", "visible_only": True,
                                       "enabled_only": True, "found_index": 0}).click()
                self.app.parent_back(1)

                self.app.find_element({"title": "", "class_name": "", "control_type": "DataGrid", "visible_only": True,
                                       "enabled_only": True, "found_index": 0}).type_keys("^s", click=True, clear=False)

                self.app.find_element(
                    {"title": "Имя файла:", "class_name": "Edit", "control_type": "Edit", "visible_only": True,
                     "enabled_only": True, "found_index": 0}).type_keys(tmp_file_path, self.app.keys.TAB, clear=True)
                self.app.find_element(
                    {"title": "Тип файла:", "class_name": "AppControlHost", "control_type": "ComboBox",
                     "visible_only": True, "enabled_only": True, "found_index": 0}).click()
                # * save it as xlsx
                self.app.find_element(
                    {"title": "Лист Excel2007-... (*.xlsx)", "class_name": "", "control_type": "ListItem",
                     "visible_only": True, "enabled_only": True, "found_index": 0}).click()
                # * click save button
                self.app.find_element(
                    {"title": "Сохранить", "class_name": "Button", "control_type": "Button", "visible_only": True,
                     "enabled_only": True, "found_index": 0}).click()

                doc_already_exists = self.app.wait_element(
                    {"title": "Подтвердить сохранение в виде", "class_name": "#32770", "control_type": "Window",
                     "visible_only": True, "enabled_only": True, "found_index": 0}, timeout=2)

                if doc_already_exists:
                    self.app.find_element(
                        {"title": "Да", "class_name": "CCPushButton", "control_type": "Button", "visible_only": True,
                         "enabled_only": True, "found_index": 0}).click()

                doc_wnd_entire = self.app.find_element(
                    {"title": "", "class_name": "", "control_type": "DataGrid", "visible_only": True,
                     "enabled_only": True, "found_index": 0}).parent(3)

                self.app.parent_switch(doc_wnd_entire)
                self.app.find_element(
                    {"title": "Закрыть", "class_name": "", "control_type": "Button", "visible_only": True,
                     "enabled_only": True, "found_index": 0}).click()

                self.app.parent_back(1)
                self.reports.append(tmp_file_path)


            else:
                logger.info("Already downloaded")

            if returns_exist:
                # * Если есть возвраты

                tmp_file_path = sTempPath.joinpath(
                    f"Возвраты оплат клиентам {protect_path(self.branch_name)} {get_report_date} {doc_num} {get_bank}.xlsx")
                # * check where we already downloaded it
                if tmp_file_path not in self.reports:

                    self.app.find_element(
                        {"title": "Возвраты оплат клиентам", "class_name": "", "control_type": "TabItem",
                         "visible_only": True, "enabled_only": True, "found_index": 0}).click()
                    self.app.parent_switch({"title": "Возвраты оплат клиентам", "class_name": "", "control_type": "Tab",
                                            "visible_only": True, "enabled_only": True, "found_index": 0})
                    self.app.find_element({"title_re": ".* Дата платежа", "class_name": "", "control_type": "Custom",
                                           "visible_only": True, "enabled_only": True, "found_index": 0}).click(
                        right=True)
                    self.app.parent_back(1)
                    self.app.find_element({"title": "Вывести список...", "class_name": "", "control_type": "MenuItem",
                                           "visible_only": True, "enabled_only": True, "found_index": 0}).click()
                    self.app.parent_switch(
                        {"title": "Вывести список", "class_name": "V8NewLocalFrameBaseWnd", "control_type": "Window",
                         "visible_only": True, "enabled_only": True, "found_index": 0})
                    self.app.find_element({"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                                           "enabled_only": True, "found_index": 0}).type_keys(self.app.keys.DOWN,
                                                                                              click=True, clear=False)
                    self.app.find_element({"title": "Табличный документ", "class_name": "", "control_type": "ListItem",
                                           "visible_only": True, "enabled_only": True, "found_index": 0}).click()
                    self.app.find_element(
                        {"title": "ОК", "class_name": "", "control_type": "Button", "visible_only": True,
                         "enabled_only": True, "found_index": 0}).click()
                    self.app.parent_back(1)

                    self.app.find_element(
                        {"title": "", "class_name": "", "control_type": "DataGrid", "visible_only": True,
                         "enabled_only": True, "found_index": 0}).type_keys("^s", click=True, clear=False)

                    # * type path where the table is stored as xlsx file

                    self.app.find_element(
                        {"title": "Имя файла:", "class_name": "Edit", "control_type": "Edit", "visible_only": True,
                         "enabled_only": True, "found_index": 0}).type_keys(tmp_file_path, self.app.keys.TAB, clear=True)
                    self.app.find_element(
                        {"title": "Тип файла:", "class_name": "AppControlHost", "control_type": "ComboBox",
                         "visible_only": True, "enabled_only": True, "found_index": 0}).click()
                    # * save it as xlsx
                    self.app.find_element(
                        {"title": "Лист Excel2007-... (*.xlsx)", "class_name": "", "control_type": "ListItem",
                         "visible_only": True, "enabled_only": True, "found_index": 0}).click()
                    # * click save button
                    self.app.find_element(
                        {"title": "Сохранить", "class_name": "Button", "control_type": "Button", "visible_only": True,
                         "enabled_only": True, "found_index": 0}).click()

                    oc_already_exists = self.app.wait_element(
                        {"title": "Подтвердить сохранение в виде", "class_name": "#32770", "control_type": "Window",
                         "visible_only": True, "enabled_only": True, "found_index": 0}, timeout=2)

                    if doc_already_exists:
                        self.app.find_element({"title": "Да", "class_name": "CCPushButton", "control_type": "Button",
                                               "visible_only": True, "enabled_only": True, "found_index": 0}).click()

                    doc_wnd_entire = self.app.find_element(
                        {"title": "", "class_name": "", "control_type": "DataGrid", "visible_only": True,
                         "enabled_only": True, "found_index": 0}).parent(3)

                    self.app.parent_switch(doc_wnd_entire)
                    self.app.find_element(
                        {"title": "Закрыть", "class_name": "", "control_type": "Button", "visible_only": True,
                         "enabled_only": True, "found_index": 0}).click()

                    self.app.parent_back(1)
                    self.reports.append(tmp_file_path)

            self.app.find_element(
                {"title": "Поступления оплат от клиентов", "class_name": "", "control_type": "TabItem",
                 "visible_only": True, "enabled_only": True, "found_index": 0}).click()

            wnd_entire = self.app.find_element(
                {"title": "Поступления оплат от клиентов", "class_name": "", "control_type": "Tab",
                 "visible_only": True, "enabled_only": True, "found_index": 0}).parent(4)
            self.app.parent_switch(wnd_entire)

            self.app.find_element({"title": "Закрыть", "class_name": "", "control_type": "Button", "visible_only": True,
                                   "enabled_only": True, "found_index": 0}).click()
            self.app.parent_back(1)
            print(f"Закончили скачивание инфы по {n_trans}")
            n_trans += 1
            trans_exists = self.app.wait_element(
                {"title_re": ".* Номер", "class_name": "", "control_type": "Custom", "visible_only": True,
                 "enabled_only": True, "found_index": n_trans}, timeout=3)

        logger.info("1с закончился успешно.")

    def edit_main_excel_file(self, beznal_z, beznal_sprut, beznal_po_vipiske):

        logger.info("Edit main excel file starts")
        wb = openpyxl.load_workbook(self.main_excel_file, data_only=False)
        print(f"sheet names of {self.main_excel_file}")
        logger.info("Прочитали файл")
        int_process_month = self.process_date.month
        int_process_year = self.process_date.year
        month_name_rus = monthes[int_process_month]
        needed_sheet_name = None
        for sheet_name in wb.sheetnames:
            if f"{month_name_rus}{int_process_year}" in sheet_name or f"{month_name_rus} {int_process_year}" in sheet_name or f"{month_name_rus}{str(int_process_year)[2:]}" in sheet_name:
                needed_sheet_name = sheet_name
                break
        if not needed_sheet_name:
            logger.info("Creating new excel list")
            needed_sheet_name = f"{month_name_rus}{int_process_year}"
            tm_wb = openpyxl.load_workbook(template_path)
            template_ws = tm_wb['Template']
            dest_ws = wb.create_sheet(needed_sheet_name)
            dest_ws.column_dimensions = template_ws.column_dimensions
            for row_num, row in enumerate(template_ws.iter_rows()):
                if row_num > 45:
                    break
                for col_num, cell in enumerate(row):
                    dest_cell = dest_ws.cell(row=row_num + 1, column=col_num + 1)
                    dest_cell.value = cell.value
                    dest_cell.number_format = cell.number_format
                    dest_cell.font = openpyxl.styles.Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        underline=cell.font.underline,
                        strike=cell.font.strike,
                        color=cell.font.color
                    )
                    dest_cell.alignment = openpyxl.styles.Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        text_rotation=cell.alignment.textRotation,
                        wrap_text=cell.alignment.wrapText,
                        shrink_to_fit=cell.alignment.shrinkToFit,
                        indent=cell.alignment.indent,
                        relativeIndent=cell.alignment.relativeIndent,
                        justifyLastLine=cell.alignment.justifyLastLine,
                        readingOrder=cell.alignment.readingOrder,
                    )
                    dest_cell.border = openpyxl.styles.Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom,
                        diagonal=cell.border.diagonal,
                        diagonal_direction=cell.border.diagonal_direction,
                        start=cell.border.start,
                        end=cell.border.end
                    )
                    dest_cell.fill = openpyxl.styles.PatternFill(
                        fill_type=cell.fill.fill_type,
                        start_color=cell.fill.start_color,
                        end_color=cell.fill.end_color
                    )
            tm_wb.close()

        logger.info(f"Нашли страницу{needed_sheet_name}")
        day: int = self.process_date.day
        index_of_process_date: int = day + 2
        wb[needed_sheet_name].cell(index_of_process_date, 2).value = self.process_date.strftime("%d.%m.%Y")
        wb[needed_sheet_name].cell(index_of_process_date, 3).value = float(beznal_z)
        wb[needed_sheet_name].cell(index_of_process_date, 4).value = float(beznal_sprut)

        # wb[needed_sheet_name].cell(index_of_process_date, 5).value = (float(beznal_z)-float(beznal_sprut))
        wb[needed_sheet_name].cell(index_of_process_date, 6).value = float(beznal_po_vipiske)
        # wb[needed_sheet_name].cell(index_of_process_date, 7).value = (float(beznal_po_vipiske)-float(beznal_sprut))

        #
        """ Безнал по Z - in column C
            Безнал по БД СПРУТ D
            Расхождение между Z и БД Спрут E
            Безнал по Выписке банка F
            Расхождение между ВБ и БД G
        """
        logger.info("Saving main excel file")
        wb.save(self.main_excel_file)
        logger.info("Edit main excel file ended")

    @retry_n_times(3)
    def get_sprut(self, report_path):
        app = Sprut(sprut_base)
        app.run()
        try:
            app.open('Отчеты')

            # * почистить Загрузки
            for path in list(Path.home().joinpath('Downloads').glob('*_Сверка_Z_отчётов_и_оборота_Спрут_*.xlsx')):
                path.unlink()
            # * Выбрать отчет Сверка Z отчётов и оборота Спрут [№100912  Flex Excel] В разработке
            app.search({"title": "", "class_name": "TvmsDBTreeList", "control_type": "Pane", "visible_only": True,
                        "enabled_only": True, "found_index": 1},
                       'Сверка Z отчётов и оборота Спрут [№100912  Flex Excel] В разработке')
            app.find_element(
                {"title": "", "class_name": "", "control_type": "SplitButton", "visible_only": True,
                 "enabled_only": True,
                 "found_index": 4}).click()
            # current_windw = app.window_element_info
            app.parent_switch(
                {"title": "N100912-Сверка Z отчётов и оборота Спрут", "class_name": "TfrmParams",
                 "control_type": "Window",
                 "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}, )
            # * прошедший рабочий день
            app.find_element(
                {"title": "", "class_name": "TcxCustomInnerTextEdit", "control_type": "Edit", "visible_only": True,
                 "enabled_only": True, "found_index": 1}).type_keys(self.process_date.strftime('%d.%m.%Y'),
                                                                    app.keys.TAB * 2, clear=True)
            # * прошедший рабочий день
            sleep(0.4)
            app.find_element(
                {"title": "", "class_name": "TcxCustomInnerTextEdit", "control_type": "Edit", "visible_only": True,
                 "enabled_only": True, "found_index": 0}).type_keys(self.process_date.strftime('%d.%m.%Y'),
                                                                    app.keys.TAB,
                                                                    clear=True)
            # * указываем один филиал из справочника филиалов в файле распределение филиалов

            result_ = app.search(
                {"class_name": "TvmsParDbLookupComboBox", "control_type": "Pane", "visible_only": True,
                 "enabled_only": True, "found_index": 0}, self.sprut_name, replace=False)
            if result_ is None:
                # TODO make screenshot
                app.quit()
                err = Exception('Ошибка в названии филиала')
                err.screen = True
                err.app = 'Спрут'
                raise err
            # * запуск выгрузки
            app.find_element(
                {"title": "Ввод", "class_name": "TvmsFooterButton", "control_type": "Button", "visible_only": True,
                 "enabled_only": True, "found_index": 0}).click()
            app.wait_element(
                {"title": "Ввод", "class_name": "TvmsFooterButton", "control_type": "Button", "visible_only": True,
                 "enabled_only": True, "found_index": 0}, until=False)
            # app.window_element_info = current_windw
            # app.parent_switch({"class_name": "TvmsProgressDlg", "control_type": "Window", "visible_only": True,
            #             "enabled_only": True, "found_index": 0, "parent" : None}, timeout=5)
            # * ожидание появления файла в Загрузках
            path = check_file_downloaded(Path.home().joinpath('Downloads', '*_Сверка_Z_отчётов_и_оборота_Спрут_*.xlsx'),
                                         timeout=3600)
            if not path:
                # todo make screenshot
                app.quit()
                err = Exception('Отчет не выгружен')
                err.screen = True
                err.app = 'Спрут'
                raise err

        except Exception as ex:
            logger.info(str(ex))
            raise ex
        finally:
            app.quit()
        # * перенос выгруженного отчета в папку обработки
        with suppress(Exception):
            move(str(path), str(report_path))
            print("should move")

    @retry_n_times(3)
    def get_sprut_101160(self, report_path):
        app = Sprut(sprut_base)
        app.run()
        try:
            app.open('Отчеты')
            # * почистить Загрузки
            for path in list(Path.home().joinpath('Downloads').glob('*_Выручка_по_кассам_с_выбором_форм_оплат_*.xlsx')):
                path.unlink()
            # * Выбрать отчет в Спруте Выручка по кассам с выбором форм оплат [101160  Flex Excel] В разработке
            app.search({"title": "", "class_name": "TvmsDBTreeList", "control_type": "Pane", "visible_only": True,
                        "enabled_only": True, "found_index": 1},
                       '%101160%')
            app.find_element(
                {"title": "", "class_name": "", "control_type": "SplitButton", "visible_only": True,
                 "enabled_only": True,
                 "found_index": 4}).click()
            # current_windw = app.window_element_info
            app.parent_switch(
                {"title": "N101160-Выручка по кассам с выбором форм оплат", "class_name": "TfrmParams",
                 "control_type": "Window", "visible_only": True, "enabled_only": True, "found_index": 0,
                 "parent": None})
            # * прошедший рабочий день
            app.find_element(
                {"title": "", "class_name": "TcxCustomInnerTextEdit", "control_type": "Edit", "visible_only": True,
                 "enabled_only": True, "found_index": 1}).type_keys(self.process_date.strftime('%d.%m.%Y'),
                                                                    app.keys.TAB * 2, clear=True)
            # * прошедший рабочий день
            sleep(0.4)
            app.find_element(
                {"title": "", "class_name": "TcxCustomInnerTextEdit", "control_type": "Edit", "visible_only": True,
                 "enabled_only": True, "found_index": 0}).type_keys(self.process_date.strftime('%d.%m.%Y'),
                                                                    app.keys.TAB,
                                                                    clear=True)
            # * указываем торг площадка из справочника филиалов в файле распределение филиалов
            stores: list = []
            if ";" in self.store_names:
                stores = self.store_names.split(";")
                logger.info(f"store_names:{stores}")
                result_ = app.search_multiple(
                    {"title": "", "class_name": "TcxCustomDropDownInnerEdit", "control_type": "Edit",
                     "visible_only": True,
                     "enabled_only": True, "found_index": 4}, stores, replace=False)
            else:
                stores = self.store_names
                result_ = app.search(
                    {"title": "", "class_name": "TcxCustomDropDownInnerEdit", "control_type": "Edit",
                     "visible_only": True,
                     "enabled_only": True, "found_index": 4}, stores, replace=False)

            if result_ is None:
                # TODO make screenshot
                app.quit()
                err = Exception('Ошибка в названии филиала')
                err.screen = True
                err.app = 'Спрут'
                raise err
            # * запуск выгрузки
            app.find_element(
                {"title": "Ввод", "class_name": "TvmsFooterButton", "control_type": "Button", "visible_only": True,
                 "enabled_only": True, "found_index": 0}).click()
            app.wait_element(
                {"title": "Ввод", "class_name": "TvmsFooterButton", "control_type": "Button", "visible_only": True,
                 "enabled_only": True, "found_index": 0}, until=False)
            # app.window_element_info = current_windw
            # app.parent_switch({"class_name": "TvmsProgressDlg", "control_type": "Window", "visible_only": True,
            #             "enabled_only": True, "found_index": 0, "parent" : None}, timeout=5)
            # * ожидание появления файла в Загрузках
            path = check_file_downloaded(
                Path.home().joinpath('Downloads\\*_Выручка_по_кассам_с_выбором_форм_оплат_*.xlsx'),
                timeout=3600)
            if not path:
                # todo make screenshot
                app.quit()
                err = Exception('Отчет не выгружен')
                err.screen = True
                err.app = 'Спрут'
                raise err
        except Exception as ex:
            logger.info(str(ex))
            raise ex
        finally:
            app.quit()

        # * перенос выгруженного отчета в папку обработки
        with suppress(Exception):
            move(str(path), str(report_path))
            print("should move")

    def read_downloaded_excel_files(self):
        columns = ['Дата транзакции', 'Номер платежной карты', 'Сумма']
        tmp_save = razbor_save_path.joinpath(f"output {self.process_date.strftime('%d.%m.%Y')}")
        tmp_save.mkdir(exist_ok=True, parents=True)
        razbor_create_path = razbor_save_path.joinpath(f"output {self.process_date.strftime('%d.%m.%Y')}",
                                                       f"Разбор {protect_path(self.branch_name)} {self.process_date.strftime('%d.%m.%Y')}.xlsx")
        main_df = pd.DataFrame(columns=columns)
        for path in self.reports:
            fix_excel_file_error(path)
            df = pd.read_excel(path, sheet_name="TDSheet")

            filtered_df = df[['Дата транзакции', 'Номер платежной карты', 'Сумма']]
            # * Тут новое требование добавить банк из 1с, название банка в названии файла
            bank = str(path)
            bank_str = ""
            if "kaspi" in str(bank).lower():
                bank_str = "Б_каспи"
            elif "народный" in str(bank).lower():
                bank_str = "Б_народный"
            else:
                bank_str = "Другой банк"

            filtered_df.loc[:, 'Банк'] = str(bank_str)

            filtered_df.loc[:, 'Дата транзакции'] = filtered_df['Дата транзакции'].astype(str)

            # filtered_df['Дата транзакции'] = filtered_df['Дата транзакции'].astype(str)
            filtered_df = filtered_df[
                filtered_df['Дата транзакции'].str.contains(self.process_date.strftime('%d.%m.%Y'))]

            if len(filtered_df) > 0:

                if "Возвраты" in str(path):
                    filtered_df['Сумма'] = filtered_df['Сумма'] * -1
                main_df = main_df._append(filtered_df, ignore_index=True)

        main_df = main_df.rename(columns={'Дата транзакции': 'Дата и время'})

        total_sum = main_df['Сумма'].sum()

        print(f"total_sum {total_sum}")
        main_df.to_excel(razbor_create_path, sheet_name="Общий", header=True, index=False)
        return total_sum, razbor_create_path

    def read_downloaded_excel_files_kzo(self):
        columns = ['Дата транзакции', 'Номер платежной карты', 'Сумма']
        tmp_save = razbor_save_path.joinpath(f"output {self.process_date.strftime('%d.%m.%Y')}")
        tmp_save.mkdir(exist_ok=True, parents=True)
        razbor_create_path = razbor_save_path.joinpath(f"output {self.process_date.strftime('%d.%m.%Y')}",
                                                       f"Разбор {protect_path(self.branch_name)} {self.process_date.strftime('%d.%m.%Y')}.xlsx")
        main_df = pd.DataFrame(columns=columns)
        for path in self.reports:
            fix_excel_file_error(path)
            df = pd.read_excel(path, sheet_name="TDSheet")
            filtered_df = df[['Дата транзакции', 'Номер платежной карты', 'Сумма']]
            bank = str(path)
            bank_str = ""
            if "kaspi" in bank.lower():
                bank_str = "Б_каспи"
            elif "народный" in bank.lower():
                bank_str = "Б_народный"
            else:
                bank_str = "Другой банк"

            filtered_df.loc[:, 'Банк'] = str(bank_str)
            # Define start and end times for filtering
            start_time = datetime.datetime.combine(self.process_date, datetime.time(hour=1))
            end_time = datetime.datetime.combine(self.process_date + datetime.timedelta(days=1), datetime.time(hour=1))

            # Filter the data based on the date/time range
            filtered_df = filtered_df[(pd.to_datetime(filtered_df['Дата транзакции'], dayfirst=True) >= start_time) & (
                    pd.to_datetime(filtered_df['Дата транзакции'], dayfirst=True) < end_time)]

            if len(filtered_df) > 0:
                print("should transfer")
                if "Возвраты" in str(path):
                    filtered_df['Сумма'] = filtered_df['Сумма'] * -1
                main_df = main_df._append(filtered_df, ignore_index=True)
            else:
                print("should not transfer")

        main_df = main_df.rename(columns={'Дата транзакции': 'Дата и время'})
        total_sum = main_df['Сумма'].sum()

        main_df.to_excel(razbor_create_path, sheet_name="Общий", header=True, index=False)
        print(f"total_sum {total_sum}")
        return total_sum, razbor_create_path

    def get_values_from_912(self, report_path):
        wb = load_workbook(report_path, data_only=True)
        ws = wb['Лист1']
        row_index = 0
        for row in ws.iter_rows(min_row=9, max_col=1):
            for cell in row:
                if cell.value == 'Итого без e-com:':
                    row_index = cell.row
                    break
        column_index = 0
        print(f"row_index: {row_index}")
        for row in ws.iter_rows(min_row=9):
            for cell in row:
                if cell.value == 'по программе Спрут':
                    column_index = cell.column
                    print(f"column_index {column_index}")
                    beznal_row_index = cell.row + 1
                    break

        for cell in ws[beznal_row_index]:
            if cell.value == "безнал":
                if cell.column < column_index:
                    col_index_z = cell.column
                else:
                    col_index_sprut = cell.column

        cell_value_z = ws.cell(row=row_index, column=col_index_z).value
        cell_value_sprut = ws.cell(row=row_index, column=col_index_sprut).value
        print("values:")
        print(cell_value_z, cell_value_sprut)
        return cell_value_z, cell_value_sprut

    def data_manipulation_101160(self, report_path, razbor_path):
        logger.info("Манипуляция с данными")
        df = pd.read_excel(report_path, header=9)
        razbor_df = pd.read_excel(razbor_path, sheet_name="Общий")

        print(f"Initial shape: {df.shape}")
        df = df.loc[df['Форма оплаты'].str.contains('Кредитная карточка')]
        filtered_df = df.loc[df['Тип розничного чека'].str.contains('Чек продажи|Чек возврата')]
        filtered_df.loc[filtered_df['Тип розничного чека'].str.contains('Чек продажи'), 'Сумма  по формам оплат'] *= -1
        print(f"Post shape: {filtered_df.shape}")
        filtered_df = filtered_df.rename(
            columns={'Дата и время чека': 'Дата и время', 'Сумма  по формам оплат': 'Сумма'})
        columns = ['Форма оплаты', 'Кассир/сотрудник', 'Тип розничного чека', 'Номер фискального регистратора',
                   'Номер чека', 'Дата и время', 'Сумма']
        filtered_df = filtered_df[columns]
        filtered_df.loc[:, 'Банк'] = "спрут"

        razbor_df[
            ['Номер фискального регистратора', 'Номер чека', 'Форма оплаты', 'Кассир/сотрудник', 'Тип розничного чека',
             'Сумма по формам оплаты']] = None

        razbor_df = razbor_df._append(filtered_df, ignore_index=True)

        print(f" Разбор после последнего добавления: {razbor_df.shape}")

        razbor_df['Дата и время'] = pd.to_datetime(razbor_df['Дата и время'], dayfirst=True)
        razbor_df['День'] = razbor_df['Дата и время'].dt.date
        razbor_df['Час'] = razbor_df['Дата и время'].dt.hour
        razbor_df['Час'] = razbor_df['Час'].astype(str)

        razbor_df['Код'] = razbor_df['Час'].astype(str) + '_' + razbor_df['Сумма'].astype(str)
        # razbor_df= razbor_df.assign(Код=)

        razbor_df = razbor_df.reindex(
            columns=['Код', 'Час', 'День', 'Банк', 'Номер платежной карты', 'Сумма', 'Дата и время',
                     'Номер фискального регистратора', 'Номер чека', 'Форма оплаты', 'Кассир/сотрудник',
                     'Тип розничного чека',
                     'Сумма по формам оплаты'])

        print("razbor_df shape: ", razbor_df.shape)
        sdf = razbor_df.copy()
        sdf['Сумма'].fillna(0, inplace=True)
        sdf['abstract'] = None

        sdf.loc[sdf['Сумма'] < 0, 'abstract'] = sdf['Сумма']
        sdf.loc[sdf['Сумма'] < 0, 'Сумма'] = 0
        sdf['abstract'] *= -1

        for index, row in sdf.iterrows():
            if row['Сумма'] in sdf['abstract'].values:
                a_index = index
                b_index = sdf.index[sdf['abstract'] == row['Сумма']].tolist()[0]
                sdf.loc[a_index, "Сумма"] = 0
                sdf.loc[b_index, "abstract"] = 0

        sdf.loc[sdf["abstract"] > 0, "Сумма"] = sdf["abstract"] * -1
        sdf = sdf[sdf['Сумма'] != 0]

        sdf.drop('abstract', axis=1, inplace=True)

        writer = pd.ExcelWriter(str(razbor_path), engine='openpyxl')

        print("before writing to excel")
        print(f"razbor shape {razbor_df.shape}")

        logger.info(f"Сводная: {sdf.shape}")
        self.comments = f"Сводная: {sdf.shape}"
        razbor_df.to_excel(writer, sheet_name="Общий", header=True, index=False)

        sdf.to_excel(writer, sheet_name="Сводная", header=True, index=False)

        writer._save()

    def update_set_status(self):
        update_executor_query = f"UPDATE ROBOT.ROBOT_SVERKA_BEZNALA_TEST SET status ='{self.status}', retry_count= {self.retry_count}, comments = '{self.comments}', error_message = '{self.errors}' WHERE id = '{self.id}' "
        print(f"update_executor_query: {update_executor_query}")
        with psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass) as conn:
            with conn.cursor() as c:
                c.execute(update_executor_query)
                conn.commit()


def performer():
    str_today = datetime.datetime.now().strftime("%d.%m.%Y")

    # str_today="27.05.2023"
    while True:

        select_one_query = f"""SELECT * FROM ROBOT.ROBOT_SVERKA_BEZNALA_TEST
         where (executor_name is NULL OR executor_name = '{ip_address}')
          AND status IN ('New','Retried')  ORDER BY CASE WHEN status = 'New' THEN 1 WHEN status = 'Retried' THEN 2 END,
          TO_DATE(process_date, 'dd.mm.yyyy')

"""

        conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
        c = conn.cursor()
        c.execute(select_one_query)
        row = c.fetchone()
        if row:
            if int(row[8]) > 2:
                update_executor_query = f"UPDATE ROBOT.ROBOT_SVERKA_BEZNALA_TEST SET executor_name ='{ip_address}', status ='Fail' WHERE id = '{row[0]}' "
                c.execute(update_executor_query)
                conn.commit()
                c.close()
                conn.close()
                continue
            else:
                update_executor_query = f"UPDATE ROBOT.ROBOT_SVERKA_BEZNALA_TEST SET executor_name ='{ip_address}' WHERE id = '{row[0]}' "
            c.execute(update_executor_query)
            conn.commit()
            c.close()
            conn.close()

            print(process)
            process(row)
            print("Finished the iteration")
        else:
            select_failed_query = f"""SELECT branch_name  FROM ROBOT.ROBOT_SVERKA_BEZNALA_TEST where substring(date_created from 1 for 10)= '{str_today}' AND (executor_name is NULL OR executor_name = '{ip_address}') AND status ='Fail' """
            with psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass) as conn:
                with conn.cursor() as c:
                    c.execute(select_failed_query)
                    failed_rows = c.fetchall()
            cc_whom = "abdi@magnum.kz;mukhtarova@magnum.kz"
            subject = "Сверка безнала>"

            if len(failed_rows) > 0:
                logger.info("Есть неудачные транзакции")

                body = "Робот отработал все филиалы кроме нижеследующих \n"
                for row in failed_rows:
                    body += f"{row} <br>\n"
                logger.info(body)

            else:
                body = f"Сверка Безнала успешно завершена"
                logger.info("Список неудачных пуст")
            send_message_by_smtp(body=body, subject=subject, to=[to_whom, cc_whom], url=smtp_host,
                                 username=smtp_author)
            logger.info("******************** ALL PROCESSES ARE FINISHED ****************")
            break


if __name__ == '__main__':
    ip_address = get_hostname()

    project_path = global_path.joinpath(f'.agent\\robot-sverka-beznala\\{ip_address}')
    log_path = project_path.joinpath(f'{sys.argv[1]}.log' if len(sys.argv) > 1 else 'dev.log')
    log_path.parent.mkdir(exist_ok=True, parents=True)
    logger = init_logger(tg_token=tg_token, chat_id='-1001659927369', log_path=log_path)

    system_info()
    logger.info("hi")
    try:
        performer()
    except Exception as ex:
        error_msg = "Сверка безнала Остановлен"
        logger.info(error_msg)
        raise ex